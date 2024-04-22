import os
import pickle
import cv2
import face_recognition
import numpy as np
import cvzone
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from firebase_admin import storage
from datetime import datetime
import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
#python image library
import threading        #Python threading allows you to have different parts of your program run concurrently and can simplify your design.
import openpyxl as op                       #the use of threading here is to ensure that the face recognition process (start_recognition) runs in a separate thread from the main Tkinter GUI thread
import openpyxl

# Connecting to the database
cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred,{'databaseURL':"https://face-attendence-b1f23-default-rtdb.firebaseio.com/",
'storageBucket':"face-attendence-b1f23.appspot.com"})

bucket = storage.bucket()


# Try different camera indices
for i in range(4):
    cap = cv2.VideoCapture(i)
    if cap.isOpened():
        print(f"Camera index {i} is opened successfully.")
        break
    else:
        print(f"Camera index {i} is not available.")

# Check if a camera is available
if not cap.isOpened():
    print("Error: No camera is available.")
    exit()

# Set the resolution
cap.set(3, 640)
cap.set(4, 480)

# Importing modes into a List
imgBackground = cv2.imread('resources/background.png')
folderModePath = 'resources/Modes'
modePathList = os.listdir(folderModePath)
imgModeList = [cv2.imread(os.path.join(folderModePath, path)) for path in modePathList]
print(f"Loaded {len(imgModeList)} modes.")

# Load the encoding file
print("Loading Encode File ...")
with open('EncodeFile.p', 'rb') as file:
    encodeListKnownWithIds = pickle.load(file)
encodeListKnown, studentIds = encodeListKnownWithIds
print("Encode File Loaded.")

modeType = 0
imgStudent = []
counter = 0
# Create an Excel workbook and set up the worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["year", "name", "total_attendance", "major", "last_attendance_time","major"])

# Dictionary to store the last attendance time for each student
last_attendance_time_dict = {}
def start_recognition():
    global imgBackground, modeType, counter, id, imgStudent

while True:
    # Capture frame-by-frame
    success, img = cap.read()

    # Resize and convert to RGB
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    # Find face locations and encodings
    faceCurFrame = face_recognition.face_locations(imgS)
    encodeCurFrame = face_recognition.face_encodings(imgS, faceCurFrame)

    # Update background image
    imgBackground[162:162 + 480, 55:55 + 640] = img
    imgBackground[44:44 + 633, 808:808 + 414] = imgModeList[modeType]

    for encodeFace, faceLoc in zip(encodeCurFrame, faceCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)

        matchIndex = np.argmin(faceDis)

        if matches[matchIndex]:
            y1, x2, y2, x1 = faceLoc
            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
            bbox = 55 + x1, 162 + y1, x2 - x1, y2 - y1
            imgBackground = cvzone.cornerRect(imgBackground, bbox, rt=0)
            id = studentIds[matchIndex]



            # Display attendance on the background
            studentInfo = db.reference(f'Students/{id}').get()
            if studentInfo:
                total_attendance = studentInfo.get('total_attendance', 0)
                cv2.putText(imgBackground, str(studentInfo['total_attendance']), (861, 125),
                            cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 1)
                cv2.putText(imgBackground, str(studentInfo['major']), (1006, 550),
                            cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 1)
                cv2.putText(imgBackground, str(id), (1006, 493),
                            cv2.FONT_HERSHEY_COMPLEX, 0.5, (255, 255, 255), 1)
                cv2.putText(imgBackground, str(studentInfo['standing']), (910, 625),
                            cv2.FONT_HERSHEY_COMPLEX, 0.6, (100, 100, 100), 1)
                cv2.putText(imgBackground, str(studentInfo['year']), (1025, 625),
                            cv2.FONT_HERSHEY_COMPLEX, 0.6, (100, 100, 100), 1)
                cv2.putText(imgBackground, str(studentInfo['starting_year']), (1125, 625),
                            cv2.FONT_HERSHEY_COMPLEX, 0.6, (100, 100, 100), 1)

                (w, h), _ = cv2.getTextSize(studentInfo['name'], cv2.FONT_HERSHEY_COMPLEX, 1, 1)
                offset = (414 - w) // 2
                cv2.putText(imgBackground, str(studentInfo['name']), (808 + offset, 445),
                            cv2.FONT_HERSHEY_COMPLEX, 1, (50, 50, 50), 1)

                blob = bucket.get_blob(f'Images/{id}.png')
                array = np.frombuffer(blob.download_as_string(), np.uint8)
                imgStudent = cv2.imdecode(array, cv2.COLOR_BGRA2BGR)
                imgBackground[175:175 + 216, 909:909 + 216] = imgStudent


                def can_mark_attendance(student_id):
                    if student_id in last_attendance_time_dict:
                        datetimeObject = last_attendance_time_dict[student_id]
                        secondsElapsed = (datetime.now() - datetimeObject).total_seconds()
                        return secondsElapsed > 30
                    else:
                        return True


                #update database of attendence

                #ref = db.reference(f'Students/{id}')
                #studentInfo['total_attendance'] +=1
                #ref.child('total_attendance').set(studentInfo['total_attendance'])


                if can_mark_attendance(id):
                    print(f"Attendance marked for ID: {id}")
                    cvzone.putTextRect(imgBackground, "Loading", (275, 400))
                    cv2.imshow("Face Attendance", imgBackground)
                    cv2.waitKey(1)
                    counter = 1
                    modeType = 1

                    # Increment attendance count
                    student_info = db.reference(f'Students/{id}').get()
                    student_info['total_attendance'] += 1

                    # Update last attendance time
                    datetime_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    student_info['last_attendance_time'] = datetime_now

                    # Save changes to Firebase
                    ref = db.reference(f'Students/{id}')
                    ref.update({'total_attendance': student_info['total_attendance'],
                                'last_attendance_time': datetime_now})

                    # Find and update corresponding row in Excel sheet
                    for row in ws.iter_rows():
                        if row[0].value == id:
                            row[2].value = student_info['total_attendance']
                            row[3].value = datetime_now
                            break

                    # Save changes to Excel file
                    wb.save("attendance_record.xlsx")

                    # Update the last attendance time in the dictionary
                    last_attendance_time_dict[id] = datetime.now()

            # Update mode
            modeType = 1

    # Display the resulting frame
    cv2.imshow("Face Attendance", imgBackground)

    # Check for 'q' key press to exit
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release video capture and close windows
cap.release()
cv2.destroyAllWindows()

def update_gui():
    _, img = cap.read()
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    img = cv2.resize(img, (640, 480))
    img = Image.fromarray(img)
    img = ImageTk.PhotoImage(image=img)

    label_webcam.imgtk = img
    label_webcam.configure(image=img)

    root.update()

    root.after(10, update_gui)

def close_face_recognition():
    global cap
    cap.release()
    root.destroy()


# Function to check if attendance can be marked for a student
def can_mark_attendance(student_id):
    if student_id in last_attendance_time_dict:
        datetimeObject = last_attendance_time_dict[student_id]
        secondsElapsed = (datetime.now() - datetimeObject).total_seconds()
        return secondsElapsed > 30
    else:
        return True
#####################################################################################


  #############################################################################

# Create a Tkinter window
root = tk.Tk()
root.title("Face Recognition Attendance System by Ali Suleman pvt.Ltd")
root.geometry("800x600")


###################################################################################
# Create a label for displaying webcam feed
label_webcam = tk.Label(root)
label_webcam.pack(padx=10, pady=10)

# Create a start button
start_button = tk.Button(root, text="Start Recognition", command=lambda: threading.Thread(target=start_recognition).start())
start_button.pack(pady=10)

# Create a close button
close_button = tk.Button(root, text="Close Face Recognition", command=close_face_recognition)
close_button.pack(pady=10)

# Function to update the GUI
root.after(10, update_gui) #10 milliseconds in loop
root.mainloop()

import openpyxl

# Create a workbook object
wb = openpyxl.Workbook()

# Access or create a worksheet within the workbook
ws = wb.active

# Write data to the worksheet
data = [
    ["major", "name", "year"],
    [1, "John", 25],
    [2, "Alice", 30],
    [3, "Bob", 28]
]

for row in data:
    ws.append(row)

# Save the workbook to a file
wb.save("attendance_record.xlsx")
